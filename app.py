#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
座位牌生成器主程序
支持批量生成可打印的座位牌PDF文件
"""

import streamlit as st
import pandas as pd
import os
from jinja2 import Template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A5, A6, letter
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import tempfile
import base64
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import os
import sys
import re
import hmac
import hashlib
import time
import json

# 隐藏Streamlit弃用警告
import warnings
warnings.filterwarnings("ignore")

# ========== 认证配置 ==========
# 共享密钥（与admin.html保持一致）
SHARED_SECRET = "seat-card-secret-key-2025"
# 登录密码
LOGIN_PASSWORD = "123456"
# Token有效时间（秒）
TOKEN_VALID_DURATION = 3600  # 1小时

def verify_token(token_str):
    """验证免密令牌"""
    try:
        # token格式: timestamp:signature
        parts = token_str.split(":")
        if len(parts) != 2:
            return False
        timestamp = int(parts[0])
        signature = parts[1]
        # 检查时间是否过期
        if time.time() - timestamp > TOKEN_VALID_DURATION:
            return False
        # 验证签名
        expected = hmac.new(
            SHARED_SECRET.encode(),
            str(timestamp).encode(),
            hashlib.sha256
        ).hexdigest()
        return hmac.compare_digest(signature, expected)
    except Exception:
        return False

def authenticate():
    """认证检查，返回是否已认证"""
    # 检查会话状态
    if st.session_state.get("authenticated", False):
        return True
    
    # 检查URL参数中的token
    params = st.query_params
    token = params.get("token", "")
    if token and verify_token(token):
        st.session_state["authenticated"] = True
        # 清除URL中的token参数
        st.query_params.clear()
        return True
    
    return False

def show_login_page():
    """显示登录页面"""
    st.markdown("""
        <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 40px;
            background: white;
            border-radius: 16px;
            box-shadow: 0 4px 24px rgba(0,0,0,0.12);
            text-align: center;
        }
        .login-container h1 {
            color: #1f77b4;
            margin-bottom: 8px;
        }
        .login-container p {
            color: #666;
            margin-bottom: 24px;
        }
        .login-container input {
            width: 100%;
            padding: 12px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 16px;
            margin-bottom: 16px;
            box-sizing: border-box;
        }
        .login-container input:focus {
            outline: none;
            border-color: #1f77b4;
        }
        .login-container button {
            width: 100%;
            padding: 12px;
            background: #1f77b4;
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            cursor: pointer;
            font-weight: bold;
        }
        .login-container button:hover {
            background: #1663a0;
        }
        .login-error {
            color: #e74c3c;
            margin-top: 12px;
            font-size: 14px;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="login-container">
            <h1>🎫 座位牌生成系统</h1>
            <p>请输入密码登录</p>
        </div>
    """, unsafe_allow_html=True)
    
    password = st.text_input("密码", type="password", key="login_password", label_visibility="collapsed")
    
    if st.button("登录", key="login_btn", use_container_width=True):
        if password == LOGIN_PASSWORD:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("密码错误，请重试")
    
    return False

# 执行认证检查
if not authenticate():
    show_login_page()
    st.stop()

# 页面配置（认证通过后）
st.set_page_config(
    page_title="座位牌生成器",
    page_icon="🎫",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 中文数字映射
cnNums = { '一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
  '十一':11,'十二':12,'十三':13,'十四':14,'十五':15,'十六':16,'十七':17,'十八':18,'十九':19,'二十':20,'二十一':21 }

# 布局关键字（不算人名）
KEYWORDS = r'^(舞台|舞台区域|投影仪|演讲台|中心线|沙发|过道|横向过道|纵向过道)$'
LAYOUT_PHRASES = r'^(课桌式|宴会厅|共计|现报名|需要增加|一排.+人|共.+排)$'

def isLayoutKeyword(s):
    """判断是否为布局关键字"""
    if not s or not isinstance(s, str):
        return True
    s = s.strip()
    if len(s) == 0:
        return True
    if re.match(KEYWORDS, s):
        return True
    if re.match(r'^第.+排$', s):
        return True
    if re.match(r'^沙发第.+排$', s):
        return True
    if re.match(LAYOUT_PHRASES, s):
        return True
    return False

def parseRowLabel(text):
    """解析排标签"""
    m = re.match(r'第(.+)排', text)
    if not m:
        return None
    return cnNums.get(m[1])

# 按列间隔分组（间隔>1列视为过道）
def groupByGap(positions):
    """按列间隔分组"""
    if not positions:
        return []
    positions.sort(key=lambda x: x['col'])
    groups = [[positions[0]]]
    for i in range(1, len(positions)):
        if positions[i]['col'] - positions[i-1]['col'] > 1:
            groups.append([])
        groups[-1].append(positions[i])
    return groups

# 从座位头构建 列号→座位号 映射表
def buildColToSeatMap(seatHeader):
    """构建列号到座位号的映射"""
    map = {}
    for g in seatHeader['groups']:
        for s in g:
            map[s['col']] = s['num']
    return map

def parseExcelLayoutFromFinder(file_path):
    """调用座位查找系统的parse-excel.js解析Excel文件，返回姓名和位置的列表（支持重名）"""
    import subprocess
    import json
    import shutil
    
    finder_dir = os.path.dirname(os.path.abspath(__file__))
    parse_script = os.path.join(finder_dir, 'parse-excel.js')
    data_json_path = os.path.join(finder_dir, 'data', 'data.json')
    
    # 确保 data 目录存在
    os.makedirs(os.path.join(finder_dir, 'data'), exist_ok=True)
    
    # 将文件复制到座位查找系统目录
    uploaded_path = os.path.join(finder_dir, 'uploaded.xlsx')
    try:
        shutil.copy(file_path, uploaded_path)
    except Exception as e:
        print(f"复制文件失败: {e}")
        return []
    
    # 调用 parse-excel.js
    try:
        result = subprocess.run(
            ['node', parse_script, uploaded_path],
            cwd=finder_dir,
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode != 0:
            print(f"parse-excel.js 执行失败: {result.stderr}")
            return []
        
        print(f"[parse-excel] 输出:\n{result.stdout}")
        
    except subprocess.TimeoutExpired:
        print("parse-excel.js 执行超时")
        return []
    except Exception as e:
        print(f"调用 parse-excel.js 失败: {e}")
        return []
    
    # 读取解析结果
    if os.path.exists(data_json_path):
        try:
            with open(data_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 返回 (name, position) 列表，支持重名
            name_position_list = []
            for attendee in data.get('attendees', []):
                if attendee.get('source') == 'excel':
                    name = attendee['name']
                    position = f"{attendee['row']} {attendee['seat']}号"
                    name_position_list.append((name, position))
            
            print(f"[parseExcelLayoutFromFinder] 提取到 {len(name_position_list)} 条记录（含重名）")
            return name_position_list
        except Exception as e:
            print(f"读取 data.json 失败: {e}")
    
    return []

def parseExcelLayout(file_path):
    """解析Excel座位布局，完全按照座位查询系统的逻辑"""
    try:
        import openpyxl
        
        name_position_list = []
        
        print(f"解析Excel文件: {file_path}")
        
        # 使用openpyxl读取，与XLSX.js保持一致
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        print(f"工作表数量: {len(wb.sheetnames)}")
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            print(f"\n处理工作表: {sheet_name}")
            
            ws = wb[sheet_name]
            
            # 读取数据，与XLSX.utils.sheet_to_json(ws, { header: 1, defval: '' }) 保持一致
            data = []
            max_row = ws.max_row
            max_col = ws.max_column
            
            for i in range(1, max_row + 1):
                row = []
                for j in range(1, max_col + 1):
                    cell = ws.cell(row=i, column=j)
                    val = cell.value
                    if val is None:
                        val = ''
                    row.append(val)
                data.append(row)
            
            print(f"工作表行数: {len(data)}")
            
            excel_attendees = []
            
            row_infos = []
            seat_headers = []
            sofa_rows = []
            
            for i in range(len(data)):
                row = data[i]
                
                # 检测沙发排
                has_sofa = any(c == '沙发' for c in row if isinstance(c, str))
                sofa_label_cells = [c for c in row if isinstance(c, str) and re.match(r'^沙发第.+排$', c.strip())]
                row_label_cells = [c for c in row if isinstance(c, str) and re.match(r'^第.+排$', c.strip())]
                row_label_cell = row_label_cells[0] if row_label_cells else None
                
                # 格式1: 同一行有"沙发" + "第X排" + 数字
                if has_sofa and row_label_cell:
                    seat_positions = []
                    for ci, c in enumerate(row):
                        if isinstance(c, (int, float)):
                            seat_positions.append({'col': ci, 'num': int(c)})
                    groups = groupByGap(seat_positions)
                    sofa_rows.append({
                        'excelRow': i,
                        'label': '沙发' + row_label_cell,
                        'rowNum': parseRowLabel(row_label_cell),
                        'groups': groups
                    })
                    continue
                
                # 格式2: 行中有"沙发第X排"
                if sofa_label_cells:
                    sofa_label = sofa_label_cells[0].strip()
                    m = re.match(r'沙发(第.+排)', sofa_label)
                    if m:
                        rn = parseRowLabel(m[1])
                        nearest_seat_header = None
                        for k in range(len(seat_headers)-1, -1, -1):
                            if seat_headers[k]['excelRow'] < i:
                                nearest_seat_header = seat_headers[k]
                                break
                        sofa_rows.append({
                            'excelRow': i,
                            'label': sofa_label,
                            'rowNum': rn,
                            'groups': nearest_seat_header['groups'] if nearest_seat_header else []
                        })
                    continue
                
                nums = []
                for ci, c in enumerate(row):
                    if isinstance(c, (int, float)):
                        nums.append({'col': ci, 'num': int(c)})
                
                # 同行既有排标签又有座位号
                if row_label_cell and len(nums) >= 3:
                    groups = groupByGap(nums)
                    seat_headers.append({'excelRow': i, 'groups': groups})
                    rn = parseRowLabel(row_label_cell)
                    if rn is not None:
                        row_infos.append({'excelRow': i, 'label': row_label_cell, 'rowNum': rn})
                    continue
                
                # 纯排标签行
                if row_label_cell:
                    rn = parseRowLabel(row_label_cell)
                    if rn is not None:
                        unique_labels = list(set([c.strip() for c in row_label_cells]))
                        if len(unique_labels) <= 1:
                            row_infos.append({'excelRow': i, 'label': row_label_cell, 'rowNum': rn})
                    continue
                
                # 纯座位编号头行
                if len(nums) >= 3 and not row_label_cell:
                    groups = groupByGap(nums)
                    seat_headers.append({'excelRow': i, 'groups': groups})
            
            # 添加沙发排
            sofa_rows = [sr for sr in sofa_rows if sr['rowNum'] is not None]
            sofa_rows.sort(key=lambda x: x['rowNum'])
            for sr_idx, sr in enumerate(sofa_rows):
                col_map = buildColToSeatMap(sr)
                end_row = len(data)
                if sr_idx < len(sofa_rows) - 1:
                    end_row = sofa_rows[sr_idx + 1]['excelRow']
                end_row = min(sr['excelRow'] + 3, end_row)
                
                for ri in range(sr['excelRow'], end_row):
                    row = data[ri]
                    for ci, cell in enumerate(row):
                        if isinstance(cell, str) and not isLayoutKeyword(cell) and ci in col_map:
                            excel_attendees.append({
                                'name': cell.strip(),
                                'row': sr['label'],
                                'seat': col_map[ci]
                            })
            
            # 为每个座位头计算过道列集合
            for sh in seat_headers:
                all_seat_cols = set()
                min_col = float('inf')
                max_col = -float('inf')
                for g in sh['groups']:
                    for s in g:
                        all_seat_cols.add(s['col'])
                        if s['col'] < min_col:
                            min_col = s['col']
                        if s['col'] > max_col:
                            max_col = s['col']
                sh['aisleCols'] = set()
                for c in range(int(min_col), int(max_col) + 1):
                    if c not in all_seat_cols:
                        sh['aisleCols'].add(c)
            
            # 查找适用于某行的座位头
            def findSeatHeader(excelRowIdx):
                best = None
                for sh in seat_headers:
                    if sh['excelRow'] <= excelRowIdx:
                        if not best or sh['excelRow'] > best['excelRow']:
                            best = sh
                return best
            
            # 关联普通排与座位头
            row_infos.sort(key=lambda x: x['rowNum'])
            
            for ri in row_infos:
                excel_row = data[ri['excelRow']]
                
                person_names_in_row = []
                for ci, cell in enumerate(excel_row):
                    if isinstance(cell, str) and not isLayoutKeyword(cell):
                        person_names_in_row.append({'col': ci, 'name': cell.strip()})
                
                if len(person_names_in_row) == 0:
                    continue
                
                person_names_in_row.sort(key=lambda x: x['col'])
                
                sh = findSeatHeader(ri['excelRow'])
                
                if sh:
                    for header_group in sh['groups']:
                        for seat_pos in header_group:
                            person = next((p for p in person_names_in_row if p['col'] == seat_pos['col']), None)
                            if person:
                                excel_attendees.append({
                                    'name': person['name'],
                                    'row': ri['label'],
                                    'seat': seat_pos['num']
                                })
                else:
                    if not person_names_in_row:
                        continue
                    
                    position_groups = [[person_names_in_row[0]]]
                    for i in range(1, len(person_names_in_row)):
                        prev_col = person_names_in_row[i-1]['col']
                        cur_col = person_names_in_row[i]['col']
                        if cur_col - prev_col > 1:
                            position_groups.append([])
                        position_groups[-1].append(person_names_in_row[i])
                    
                    seat_num = 1
                    for group in position_groups:
                        min_col = min(p['col'] for p in group)
                        max_col = max(p['col'] for p in group)
                        for col in range(min_col, max_col + 1):
                            person = next((p for p in group if p['col'] == col), None)
                            if person:
                                excel_attendees.append({
                                    'name': person['name'],
                                    'row': ri['label'],
                                    'seat': seat_num
                                })
                            seat_num += 1
            
            # === U型会议室检测: 方法1 ===
            # 始终尝试U型布局解析，不管有没有找到其他记录
            if True:
                u_column_header_row = -1
                u_column_headers = []
                u_seat_num_cols = []
                
                for i in range(len(data)):
                    row = data[i]
                    col_labels = []
                    seat_label_cols = []
                    for ci, c in enumerate(row):
                        if isinstance(c, str):
                            trimmed = c.strip()
                            if re.match(r'^第.+列$', trimmed):
                                col_labels.append({'col': ci, 'label': trimmed})
                            if trimmed == '座位号':
                                seat_label_cols.append(ci)
                    if len(col_labels) >= 2:
                        u_column_header_row = i
                        u_column_headers = sorted(col_labels, key=lambda x: x['col'])
                        u_seat_num_cols = sorted(seat_label_cols)
                        break
                
                if len(u_column_headers) >= 2:
                    left_seat_col = u_seat_num_cols[0] if len(u_seat_num_cols) >= 2 else -1
                    right_seat_col = u_seat_num_cols[-1] if len(u_seat_num_cols) >= 2 else -1
                    
                    mid_idx = len(u_column_headers) // 2
                    left_cols = u_column_headers[:mid_idx]
                    right_cols = u_column_headers[mid_idx:]
                    
                    inner_left_col = left_cols[-1] if left_cols else None
                    inner_right_col = right_cols[0] if right_cols else None
                    
                    bottom_nums = []
                    bottom_row_idx = -1
                    bottom_col_map = {}
                    
                    # 查找底部行
                    for i in range(u_column_header_row + 1, len(data)):
                        row = data[i]
                        if not row:
                            continue
                        nums = []
                        for ci, c in enumerate(row):
                            if isinstance(c, (int, float)):
                                nums.append({'col': ci, 'num': int(c)})
                        
                        inner_bottom_nums = []
                        if inner_left_col and inner_right_col:
                            inner_bottom_nums = [n for n in nums if n['col'] > inner_left_col['col'] and n['col'] < inner_right_col['col']]
                        
                        if len(inner_bottom_nums) >= 2:
                            inner_bottom_nums.sort(key=lambda x: x['col'])
                            bottom_nums = [n['num'] for n in inner_bottom_nums]
                            bottom_row_idx = i
                            for n in inner_bottom_nums:
                                bottom_col_map[n['col']] = n['num']
                            break
                    
                    paired_rows = []
                    stop_row = bottom_row_idx if bottom_row_idx >= 0 else len(data)
                    for i in range(u_column_header_row + 1, stop_row):
                        row = data[i]
                        left_val = row[left_seat_col] if left_seat_col >= 0 else None
                        right_val = row[right_seat_col] if right_seat_col >= 0 else None
                        if isinstance(left_val, (int, float)) and isinstance(right_val, (int, float)):
                            paired_rows.append({'excelRow': i, 'leftNum': left_val, 'rightNum': right_val})
                    
                    # 提取人名 - 左侧所有列
                    for left_col in left_cols:
                        for i in range(u_column_header_row + 1, stop_row):
                            row = data[i]
                            if left_col['col'] < len(row) and left_seat_col < len(row):
                                # 座位号在座位号列，人名在列标签列！
                                seat_val = row[left_seat_col]
                                name_val = row[left_col['col']]
                                if isinstance(seat_val, (int, float)) and isinstance(name_val, str) and name_val.strip() and not isLayoutKeyword(name_val):
                                    excel_attendees.append({
                                        'name': name_val.strip(),
                                        'row': left_col['label'],
                                        'seat': seat_val
                                    })
                    
                    # 提取人名 - 右侧所有列
                    for right_col in right_cols:
                        for i in range(u_column_header_row + 1, stop_row):
                            row = data[i]
                            if right_col['col'] < len(row) and right_seat_col < len(row):
                                # 座位号在座位号列，人名在列标签列！
                                seat_val = row[right_seat_col]
                                name_val = row[right_col['col']]
                                if isinstance(seat_val, (int, float)) and isinstance(name_val, str) and name_val.strip() and not isLayoutKeyword(name_val):
                                    excel_attendees.append({
                                        'name': name_val.strip(),
                                        'row': right_col['label'],
                                        'seat': seat_val
                                    })
                    
                    # 提取人名 - 底部
                    if bottom_row_idx >= 0:
                        for ri in range(bottom_row_idx - 1, min(bottom_row_idx + 3, len(data))):
                            if ri < 0 or ri == bottom_row_idx:
                                continue
                            name_row = data[ri]
                            if not name_row:
                                continue
                            for ci, cell in enumerate(name_row):
                                if isinstance(cell, str) and cell.strip() and not isLayoutKeyword(cell) and ci in bottom_col_map:
                                    excel_attendees.append({
                                        'name': cell.strip(),
                                        'row': '底部',
                                        'seat': bottom_col_map[ci]
                                    })
            
            # 转换为最终结果格式
            for attendee in excel_attendees:
                name = attendee['name']
                position = f"{attendee['row']} {attendee['seat']}号"
                name_position_list.append((name, position))
                print(f"  找到: {name} -> {position}")
        
        wb.close()
        
        print(f"\n最终提取的位置信息: {len(name_position_list)} 条记录")
        return name_position_list
    except Exception as e:
        print(f"解析Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return []

def load_attendees_from_finder():
    """从座位查询系统的 data.json 读取已识别的参会者数据"""
    finder_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'data.json')
    
    if not os.path.exists(finder_data_path):
        return None, "未找到座位查询系统数据文件，请先通过座位查询系统的管理后台导入座位数据"
    
    try:
        with open(finder_data_path, 'r', encoding='utf-8') as f:
            finder_data = json.load(f)
        
        attendees = finder_data.get('attendees', [])
        venues = finder_data.get('venues', [])
        
        if not attendees:
            return None, "座位查询系统中暂无参会者数据"
        
        # 获取会场ID集合
        valid_venue_ids = set(v['id'] for v in venues)
        
        # 获取会场信息
        venue_names = {}
        for v in venues:
            venue_names[v['id']] = v.get('name', v['id'])
        
        # 按会场分组，只处理有效的会场ID
        venue_attendees = {}
        for a in attendees:
            vid = a.get('venueId', 'unknown')
            # 只保留有效的会场ID，或者将无效的会场ID映射到第一个有效会场
            if vid not in valid_venue_ids:
                if valid_venue_ids:
                    vid = next(iter(valid_venue_ids))
                else:
                    continue
            if vid not in venue_attendees:
                venue_attendees[vid] = []
            venue_attendees[vid].append(a)
        
        # 确保只返回有效的会场
        valid_venues = [v for v in venues if v['id'] in venue_attendees]
        
        return {
            'attendees': attendees,
            'venues': valid_venues,
            'venue_names': venue_names,
            'venue_attendees': venue_attendees
        }, None
    except Exception as e:
        return None, f"读取座位查询系统数据失败: {e}"

# 字符预处理：替换特殊字符为可显示字符
def preprocess_special_chars(text):
    """预处理文本中的特殊字符，确保能在PDF中正常显示"""
    if not text:
        return text
    
    # 打印原始字符的Unicode编码用于调试
    print(f"[字符预处理] 输入: '{text}'")
    for i, char in enumerate(text):
        print(f"  字符 {i}: '{char}' (U+{ord(char):04X})")
    
    # 特殊字符映射表 - 尝试用居中的间隔符
    char_mapping = {
        # 间隔号、点号类 - 使用位置较好的字符
        '\u00B7': '-',      # MIDDLE DOT → 连字符（位置居中）
        '\u2022': '-',      # BULLET
        '\u2219': '-',      # DOT OPERATOR
        '\u22C5': '-',      # DOT OPERATOR
        '\u30FB': '-',      # KATAKANA MIDDLE DOT (U+30FB)
        '·': '-',           # 任何点号都替换为连字符
        '•': '-',
        
        # 全角符号转半角
        '．': '-',
        '。': '-',
        '，': ',',
        '、': ',',
        '；': ';',
        '：': ':',
        '？': '?',
        '！': '!',
        
        # 空格类
        '\u00A0': ' ',
        '\u2000': ' ',
        '\u2001': ' ',
        '\u2002': ' ',
        '\u2003': ' ',
        '\u2004': ' ',
        '\u2005': ' ',
        '\u2006': ' ',
        '\u2007': ' ',
        '\u2008': ' ',
        '\u2009': ' ',
        '\u200A': ' ',
        '\u202F': ' ',
        '\u205F': ' ',
        '\u3000': ' ',
        
        # 其他常见特殊字符
        '—': '-',
        '–': '-',
        '―': '-',
        '…': '...',
        '‘': "'",
        '’': "'",
        '“': '"',
        '”': '"',
    }
    
    # 应用映射
    result = []
    for char in text:
        if char in char_mapping:
            mapped_char = char_mapping[char]
            result.append(mapped_char)
            if char != mapped_char:
                print(f"  映射: '{char}' (U+{ord(char):04X}) → '{mapped_char}' (U+{ord(mapped_char):04X})")
        else:
            # 检查是否是基本ASCII或CJK字符
            if ord(char) < 128:
                # ASCII 字符，直接保留
                result.append(char)
            elif 0x4E00 <= ord(char) <= 0x9FFF:
                # CJK 统一汉字，保留
                result.append(char)
            elif 0x3400 <= ord(char) <= 0x4DBF:
                # CJK 扩展 A，保留
                result.append(char)
            elif 0x20000 <= ord(char) <= 0x2A6DF:
                # CJK 扩展 B，保留
                result.append(char)
            else:
                # 其他未知字符 - 激进替换！
                print(f"  警告: 未知字符 '{char}' (U+{ord(char):04X})，替换为 '-'")
                result.append('-')
    
    final_text = ''.join(result)
    if final_text != text:
        print(f"[字符预处理] 原始: {repr(text)} → 处理后: {repr(final_text)}")
    
    return final_text

# 加载中文字体
def load_chinese_font():
    """加载中文字体"""
    font_paths = []
    
    # 优先检查本地字体目录
    local_fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
    if os.path.exists(local_fonts_dir):
        try:
            for filename in os.listdir(local_fonts_dir):
                if filename.lower().endswith(('.ttf', '.otf', '.ttc')):
                    font_paths.append(os.path.join(local_fonts_dir, filename))
        except Exception as e:
            pass  # 静默失败，继续尝试系统字体
    
    # Windows系统字体路径
    if sys.platform == 'win32':
        font_paths.extend([
            'C:\\Windows\\Fonts\\msyh.ttf',    # 微软雅黑（支持最多特殊字符）
            'C:\\Windows\\Fonts\\simhei.ttf',  # 黑体
            'C:\\Windows\\Fonts\\simsun.ttc',  # 宋体
        ])
    # Linux/Docker系统字体路径
    else:
        # DejaVu 和 Noto Sans CJK 支持最广的 Unicode 字符（优先使用）
        font_paths.extend([
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',            # DejaVu Sans - 支持最广泛的 Unicode
            '/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf',  # DejaVu Sans Condensed
            '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',     # Noto Sans CJK (OpenType) - subfontIndex=2 为简体中文
            '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',     # Noto Sans CJK (TrueType)
            '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',  # Droid Sans Fallback
            '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',               # 文泉驿正黑
            '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',             # 文泉驿微米黑
            '/usr/share/fonts/truetype/arphic/uming.ttc',                 # AR PL UMing
            '/usr/share/fonts/truetype/arphic/ukai.ttc',                  # AR PL UKai
        ])
    
    # 尝试加载字体
    for font_path in font_paths:
        if not os.path.exists(font_path):
            continue
        
        try:
            # 检查文件可读性
            if not os.access(font_path, os.R_OK):
                continue
                
            font_name = os.path.splitext(os.path.basename(font_path))[0]
            # 对于TTC文件，尝试多个子字体索引
            if font_path.lower().endswith('.ttc'):
                # 尝试多个子字体索引，避免单个索引失败
                for subfont_idx in [0, 1, 2, 3]:
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, font_path, subfontIndex=subfont_idx))
                        return font_name
                    except Exception:
                        continue
            else:
                # 普通TTF/OTF文件
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                return font_name
        except Exception as e:
            continue  # 静默失败，尝试下一个字体
    
    # 如果没有找到中文字体，返回默认字体
    return "Helvetica"

# 创建必要的目录
os.makedirs("templates", exist_ok=True)
os.makedirs("output", exist_ok=True)
os.makedirs("fonts", exist_ok=True)

def create_namecard_template():
    """创建默认的座位牌HTML模板"""
    template_content = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {
            size: {{ page_size }};
            margin: 0;
        }
        
        body {
            margin: 0;
            padding: 0;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            background-color: {{ background_color }};
            {% if background_image %}
            background-image: url('{{ background_image }}');
            background-size: cover;
            background-position: center;
            {% endif %}
        }
        
        .namecard {
            width: {{ card_width }}mm;
            height: {{ card_height }}mm;
            border: {{ border_width }}px {{ border_style }} {{ border_color }};
            display: flex;
            justify-content: center;
            align-items: center;
            background: {{ card_background }};
            box-shadow: 2px 2px 10px rgba(0,0,0,0.1);
        }
        
        .name {
            font-family: {{ font_family }};
            font-size: {{ font_size }}px;
            font-weight: {{ font_weight }};
            color: {{ font_color }};
            text-align: center;
            line-height: 1.2;
        }
    </style>
</head>
<body>
    <div class="namecard">
        <div class="name">{{ name }}</div>
    </div>
</body>
</html>
"""
    
    with open("templates/default.html", "w", encoding="utf-8") as f:
        f.write(template_content)

def load_template():
    """加载HTML模板"""
    if not os.path.exists("templates/default.html"):
        create_namecard_template()
    
    with open("templates/default.html", "r", encoding="utf-8") as f:
        return Template(f.read())

def draw_string_with_spacing(c, x, y, text, char_spacing=0):
    """绘制带字间距的文字"""
    if char_spacing == 0:
        c.drawCentredString(x, y, text)
        return
    
    # 计算每个字符的宽度和总宽度
    total_width = 0
    char_widths = []
    for char in text:
        width = c.stringWidth(char, c._fontname, c._fontsize)
        char_widths.append(width)
        total_width += width
    
    # 加上字间距
    total_width += char_spacing * (len(text) - 1)
    
    # 计算起始位置（居中）
    start_x = x - total_width / 2
    
    # 逐个绘制字符
    current_x = start_x
    for i, char in enumerate(text):
        c.drawString(current_x, y, char)
        if i < len(text) - 1:
            current_x += char_widths[i] + char_spacing

def draw_namecard(c, x, y, card_width, card_height, name, template_vars, image_path, custom_font_name, chinese_font, is_mirror, position=None):
    """绘制单个座位牌"""
    # 预处理特殊字符
    name = preprocess_special_chars(name)
    if position:
        position = preprocess_special_chars(position)
    
    # 获取座位号设置
    show_seat_number = template_vars.get('show_seat_number', True)
    seat_number_font_size = template_vars.get('seat_number_font_size', 6)
    seat_number_color = template_vars.get('seat_number_color', '#666666')
    
    # 保存当前绘图状态
    c.saveState()
    
    # 如果需要镜像显示（整体上下镜像）
    if is_mirror:
        c.translate(x, y)
        c.rotate(180)
        c.translate(-x, -y)
    
    # 绘制卡片背景
    try:
        bg_color = HexColor(template_vars['background_color'])
        c.setFillColor(bg_color)
        c.rect(x - card_width/2, y - card_height/2, card_width, card_height, fill=1)
    except:
        c.setFillColorRGB(1, 1, 1)  # 默认白色
        c.rect(x - card_width/2, y - card_height/2, card_width, card_height, fill=1)
    
    # 绘制背景图片
    if image_path:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(image_path)
            # 计算图片缩放比例
            img_width, img_height = img.getSize()
            scale_x = card_width / img_width
            scale_y = card_height / img_height
            scale = min(scale_x, scale_y)
            
            # 计算图片位置（居中）
            img_scaled_width = img_width * scale
            img_scaled_height = img_height * scale
            img_x = x - img_scaled_width / 2
            img_y = y - img_scaled_height / 2
            
            # 绘制图片
            c.drawImage(img, img_x, img_y, width=img_scaled_width, height=img_scaled_height)
        except Exception as e:
            print(f"绘制背景图片失败: {e}")
    
    # 绘制边框
    if template_vars['border_width'] > 0:
        try:
            border_color = HexColor(template_vars['border_color'])
            c.setStrokeColor(border_color)
        except:
            c.setStrokeColorRGB(0.8, 0.8, 0.8)  # 默认灰色
        c.setLineWidth(template_vars['border_width'])
        c.rect(x - card_width/2, y - card_height/2, card_width, card_height, stroke=1)
    
    # 设置字体和颜色
    try:
        font_color = HexColor(template_vars['font_color'])
        c.setFillColor(font_color)
    except:
        c.setFillColorRGB(0, 0, 0)  # 默认黑色
    
    # 选择字体：优先使用自定义字体，然后使用系统字体
    font_name = 'Helvetica'  # 默认字体作为最终回退
    
    # 检查自定义字体是否已注册
    if custom_font_name:
        try:
            # 尝试设置自定义字体
            c.setFont(custom_font_name, template_vars['font_size'])
            font_name = custom_font_name
        except Exception:
            # 自定义字体失败，尝试中文字体
            custom_font_name = None
    
    # 如果没有自定义字体，尝试中文字体
    if not custom_font_name and chinese_font != 'Helvetica':
        try:
            c.setFont(chinese_font, template_vars['font_size'])
            font_name = chinese_font
        except Exception:
            # 中文字体也失败，使用默认Helvetica
            font_name = 'Helvetica'
            c.setFont(font_name, template_vars['font_size'])
    elif not custom_font_name:
        # 已经是默认Helvetica
        c.setFont(font_name, template_vars['font_size'])
    
    # 处理字体粗细（仅对Helvetica有效）
    if template_vars['font_weight'] == 'bold' and font_name == 'Helvetica':
        try:
            font_name = 'Helvetica-Bold'
            c.setFont(font_name, template_vars['font_size'])
        except Exception:
            # Helvetica-Bold不可用，继续使用普通Helvetica
            pass
    
    # 获取字间距
    char_spacing = template_vars.get('char_spacing', 0)
    
    # 获取姓名位置偏移（转换为点）
    text_offset_x = template_vars.get('text_offset_x', 0) * mm
    text_offset_y = template_vars.get('text_offset_y', 0) * mm
    
    # 计算实际文字位置
    text_x = x + text_offset_x
    text_y = y + text_offset_y
    
    # 绘制姓名（支持效果叠加）
    text_effects = template_vars.get('text_effects', [])
    
    # 保存当前状态
    c.saveState()
    
    # 先设置基础文字颜色
    try:
        font_color = HexColor(template_vars['font_color'])
        c.setFillColor(font_color)
    except:
        c.setFillColorRGB(0, 0, 0)  # 默认黑色
    
    # 处理立体效果（最底层）
    if "立体" in text_effects:
        stereo_offset = template_vars.get('stereo_offset', 1.0)
        stereo_color = template_vars.get('stereo_color', '#666666')
        
        # 绘制立体效果的底层文字
        try:
            stereo_color_obj = HexColor(stereo_color)
            c.setFillColor(stereo_color_obj)
        except:
            c.setFillColorRGB(0.4, 0.4, 0.4)  # 默认灰色
        
        draw_string_with_spacing(c, text_x + stereo_offset, text_y - stereo_offset, name, char_spacing)
        draw_string_with_spacing(c, text_x + stereo_offset, text_y + stereo_offset, name, char_spacing)
        draw_string_with_spacing(c, text_x - stereo_offset, text_y - stereo_offset, name, char_spacing)
        draw_string_with_spacing(c, text_x - stereo_offset, text_y + stereo_offset, name, char_spacing)
        
        # 恢复文字颜色
        try:
            font_color = HexColor(template_vars['font_color'])
            c.setFillColor(font_color)
        except:
            c.setFillColorRGB(0, 0, 0)  # 默认黑色
    
    # 处理阴影效果（中间层）
    if "阴影" in text_effects:
        shadow_offset = template_vars.get('shadow_offset', 2.0)
        shadow_color = template_vars.get('shadow_color', '#888888')
        
        try:
            shadow_color_obj = HexColor(shadow_color)
            c.setFillColor(shadow_color_obj)
        except:
            c.setFillColorRGB(0.5, 0.5, 0.5)  # 默认灰色
        
        # 绘制阴影文字
        draw_string_with_spacing(c, text_x + shadow_offset, text_y - shadow_offset, name, char_spacing)
        
        # 恢复文字颜色
        try:
            font_color = HexColor(template_vars['font_color'])
            c.setFillColor(font_color)
        except:
            c.setFillColorRGB(0, 0, 0)  # 默认黑色
    
    # 处理勾边效果（最上层）
    if "勾边" in text_effects:
        stroke_width = template_vars.get('stroke_width', 1.0)
        stroke_color = template_vars.get('stroke_color', '#000000')
        
        # 绘制多个方向的描边，形成勾边效果
        directions = [
            (-stroke_width, -stroke_width),
            (stroke_width, -stroke_width),
            (-stroke_width, stroke_width),
            (stroke_width, stroke_width)
        ]
        
        try:
            # 设置勾边颜色
            stroke_color_obj = HexColor(stroke_color)
            c.setFillColor(stroke_color_obj)
        except:
            c.setFillColorRGB(0, 0, 0)  # 默认黑色
        
        # 绘制多个方向的描边
        for dx, dy in directions:
            draw_string_with_spacing(c, text_x + dx, text_y + dy, name, char_spacing)
        
        # 恢复原始文字颜色
        try:
            font_color = HexColor(template_vars['font_color'])
            c.setFillColor(font_color)
        except:
            c.setFillColorRGB(0, 0, 0)  # 默认黑色
    
    # 绘制主文字
    draw_string_with_spacing(c, text_x, text_y, name, char_spacing)
    
    # 恢复文字效果状态
    c.restoreState()
    
    # 绘制位置信息（小字）
    if position and show_seat_number:
        # 设置位置信息字体大小
        c.setFont(font_name, seat_number_font_size)
        # 设置座位号颜色
        try:
            seat_color = HexColor(seat_number_color)
            c.setFillColor(seat_color)
        except:
            c.setFillColorRGB(0.4, 0.4, 0.4)  # 默认灰色
        
        # 绘制位置信息在左下角底部
        position_x = x - card_width / 2 + 5  # 左边留5mm边距
        position_y = y - card_height / 2 + 5  # 底部留5mm边距
        c.drawString(position_x, position_y, position)
    
    # 恢复最外层状态
    c.restoreState()

def generate_pdf(names, template_vars, output_path):
    """为姓名列表生成PDF座位牌"""
    # 导入必要的模块
    import os
    
    # 获取座位牌类型
    card_type = template_vars.get('card_type', '普通座位牌')
    
    # 设置卡片尺寸（转换为点）
    card_width = template_vars['card_width'] * mm
    card_height = template_vars['card_height'] * mm
    
    # 获取页面尺寸
    if card_type == '三角立式台卡':
        # 三角立式台卡：设置自定义页面尺寸，左右不留空白
        paste_height = template_vars.get('paste_area_height', 48.5) * mm  # 粘贴区域高度
        total_card_height = card_height * 2 + paste_height * 2  # 总高度
        page_size = (card_width, total_card_height)  # 左右不留空白
    elif 'page_width' in template_vars and 'page_height' in template_vars:
        # 使用自定义页面尺寸（转换为点）
        page_size = (template_vars['page_width'] * mm, template_vars['page_height'] * mm)
    else:
        # 使用预设页面尺寸
        page_sizes = {
            'A3': (841.89, 1190.55),  # A3尺寸（点）
            'A4': A4,
            'A5': A5, 
            'A6': A6,
            'Letter': letter
        }
        page_size = page_sizes.get(template_vars['page_size'], A4)
    
    # 创建PDF画布
    c = canvas.Canvas(output_path, pagesize=page_size)
    
    # 加载中文字体
    chinese_font = load_chinese_font()
    
    # 处理字体（优先使用挂载字体，然后是上传字体）
    selected_system_font = template_vars.get('selected_system_font')
    font_files = template_vars.get('font_files', {})
    custom_font = template_vars.get('custom_font')
    custom_font_path = template_vars.get('custom_font_path')  # 使用已保存的临时文件路径
    custom_font_name = template_vars.get('custom_font_name')  # 使用已保存的字体名称
    
    if selected_system_font and selected_system_font in font_files:
        # 使用挂载目录中的字体
        font_path = font_files[selected_system_font]
        custom_font_name = selected_system_font
        try:
            # 处理TTC文件，尝试多个子字体索引
            if font_path.lower().endswith('.ttc'):
                for subfont_idx in [0, 1, 2, 3]:
                    try:
                        pdfmetrics.registerFont(TTFont(custom_font_name, font_path, subfontIndex=subfont_idx))
                        break
                    except Exception:
                        continue
            else:
                pdfmetrics.registerFont(TTFont(custom_font_name, font_path))
        except Exception as e:
            custom_font_name = None
    elif custom_font_path and custom_font_name:
        # 使用已保存的自定义字体临时文件
        try:
            pdfmetrics.registerFont(TTFont(custom_font_name, custom_font_path))
        except Exception as e:
            custom_font_name = None
    elif custom_font:
        # 保存上传的字体到临时文件（作为备选方案）
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.' + custom_font.name.split('.')[-1]) as temp_file:
                temp_file.write(custom_font.getbuffer())
                custom_font_path = temp_file.name
                custom_font_name = os.path.splitext(os.path.basename(custom_font.name))[0]
            
            # 注册自定义字体
            pdfmetrics.registerFont(TTFont(custom_font_name, custom_font_path))
        except Exception as e:
            custom_font_name = None
    
    # 处理背景图片
    background_image = template_vars.get('background_image')
    image_path = template_vars.get('background_image_path')  # 使用已保存的临时文件路径
    
    if not image_path and background_image:
        # 保存上传的图片到临时文件（作为备选方案）
        with tempfile.NamedTemporaryFile(delete=False, suffix='.' + background_image.name.split('.')[-1]) as temp_file:
            temp_file.write(background_image.getbuffer())
            image_path = temp_file.name
    
    # 三角立式台卡的特殊设置
    if card_type == '三角立式台卡':
        total_card_height = card_height * 2 + paste_height * 2  # 总高度（上下各有粘贴空白，中间没有）
    
    # 获取位置信息列表 [(name, position), ...]，支持重名
    name_position_list = template_vars.get('name_position_list', [])
    
    # 计算每页可以容纳的卡片数量
    page_width, page_height = page_size
    if card_type == '三角立式台卡':
        # 三角立式台卡：每个姓名一个大卡片，包含上下两部分
        # 因为已经设置了自定义页面尺寸，每页只放一个卡片
        effective_card_width = card_width
        effective_card_height = total_card_height
        cards_per_row = 1
        cards_per_col = 1
    else:
        # 普通座位牌：每个姓名一个卡片
        effective_card_width = card_width
        effective_card_height = card_height
        # 避免除以零的情况
        if effective_card_width > 0 and effective_card_height > 0:
            cards_per_row = max(1, int(page_width // effective_card_width))
            cards_per_col = max(1, int(page_height // effective_card_height))
        else:
            cards_per_row = 1
            cards_per_col = 1
    cards_per_page = cards_per_row * cards_per_col
    
    # 生成所有座位牌
    for i, name in enumerate(names):
        # 计算当前卡片的位置
        page_num = i // cards_per_page
        pos_in_page = i % cards_per_page
        
        # 如果开始新页面，先结束当前页面
        if pos_in_page == 0 and i > 0:
            c.showPage()
        
        # 计算卡片在页面中的位置
        row = pos_in_page // cards_per_row
        col = pos_in_page % cards_per_row
        
        # 获取位置信息（通过索引查找，支持重名）
        if i < len(name_position_list):
            position = name_position_list[i][1]  # (name, position) 的第二个元素
        else:
            position = None
        
        if card_type == '三角立式台卡':
            # 三角立式台卡：一个大卡片包含上下两部分
            # 从页面最左边开始绘制，占满整个页面宽度
            x = page_width / 2  # 使用页面宽度的中心
            
            # 计算总高度并垂直居中
            total_card_height = paste_height * 2 + card_height * 2
            big_card_top_y = (page_height - total_card_height) / 2  # 垂直居中
            
            # 整个大卡片的结构（从顶部到底部）：
            # ┌─────────────────────────┐  ← big_card_top_y
            # │   粘贴空白             │  ← 填充颜色
            # ├─────────────────────────┤
            # │   上半部分（正常）    │
            # │         底边 ────────┤  ← 重合处
            # ├─────────────────────────┤  ← 折叠线
            # │         顶边 ────────┤  ← 重合处
            # │   下半部分（镜像）    │
            # ├─────────────────────────┤
            # │   粘贴空白             │  ← 填充颜色
            # └─────────────────────────┘  ← big_card_top_y + total_card_height
            
            # 计算两个座位牌的中心位置
            # 上半部分（正常）的中心 - 从页面顶部往下：粘贴空白 + 座位牌中心
            top_center_y = big_card_top_y + paste_height + card_height/2
            # 下半部分（镜像）的中心 - 从页面顶部往下：粘贴空白 + 上座位牌 + 座位牌中心
            # 上半部分的底边和下半部分的顶边直接重合，中间不留白
            bottom_center_y = big_card_top_y + paste_height + card_height + card_height/2
            # 折叠线在两部分的重合处
            fold_line_y = big_card_top_y + paste_height + card_height
            
            # 填充上下空白粘贴区
            try:
                paste_color = HexColor(template_vars.get('paste_area_color', '#ffffff'))
                c.setFillColor(paste_color)
            except:
                c.setFillColorRGB(1, 1, 1)  # 默认白色
            
            # 填充上粘贴区
            c.rect(0, big_card_top_y, page_width, paste_height, fill=1)
            
            # 填充下粘贴区
            bottom_paste_top_y = big_card_top_y + paste_height + card_height * 2
            c.rect(0, bottom_paste_top_y, page_width, paste_height, fill=1)
            
            # 绘制上半部分（正常显示）
            draw_namecard(c, x, top_center_y, card_width, card_height, name, template_vars, image_path, custom_font_name, chinese_font, False)
            
            # 绘制下半部分（镜像显示，折叠后从另一面看是正的）
            draw_namecard(c, x, bottom_center_y, card_width, card_height, name, template_vars, image_path, custom_font_name, chinese_font, True)
            
            # 绘制折叠线（虚线）- 在两部分的交界处
            c.setStrokeColorRGB(0.5, 0.5, 0.5)
            c.setLineWidth(0.5)
            c.setDash(5, 2)
            c.line(0, fold_line_y, page_width, fold_line_y)
            c.setDash()
            
            # 在上下空白区域添加座位信息（折叠后会被隐藏）
            if position and template_vars.get('show_seat_number', True):
                # 设置位置信息字体
                if custom_font_name:
                    font_name = custom_font_name
                else:
                    font_name = chinese_font
                c.setFont(font_name, template_vars.get('seat_number_font_size', 6))
                # 设置座位号颜色
                try:
                    seat_color = HexColor(template_vars.get('seat_number_color', '#666666'))
                    c.setFillColor(seat_color)
                except:
                    c.setFillColorRGB(0.4, 0.4, 0.4)  # 默认灰色
                
                # 在上空白区域添加座位信息（顶部）
                top_paste_top_y = big_card_top_y + 10  # 顶部留10mm边距
                c.drawCentredString(x, top_paste_top_y, position)
                
                # 在下空白区域添加座位信息（底部）
                bottom_paste_bottom_y = big_card_top_y + paste_height + card_height * 2 + paste_height - 10  # 底部留10mm边距
                c.drawCentredString(x, bottom_paste_bottom_y, position)
        else:
            # 普通座位牌：单个卡片
            x = col * effective_card_width + effective_card_width / 2
            y = page_height - (row + 1) * effective_card_height + effective_card_height / 2
            draw_namecard(c, x, y, card_width, card_height, name, template_vars, image_path, custom_font_name, chinese_font, False, position)
    
    # 清理临时文件
    if image_path:
        try:
            import os
            os.unlink(image_path)
        except:
            pass
    
    # 清理自定义字体临时文件
    if custom_font_path:
        try:
            import os
            os.unlink(custom_font_path)
        except:
            pass
    
    # 保存PDF
    c.save()

def main():
    """主程序"""
    st.title("🎫 座位牌生成器")
    st.markdown("批量生成可打印的座位牌PDF文件")
    
    # 侧边栏 - 样式设置
    st.sidebar.header("🎨 样式设置")
    
    # 基本设置
    col1, col2 = st.sidebar.columns(2)
    with col1:
        card_width = st.number_input("座位牌宽度(mm)", min_value=1, value=210)
        font_size = st.number_input("字体大小(px)", min_value=20, max_value=100, value=60)
    with col2:
        card_height = st.number_input("座位牌高度(mm)", min_value=1, value=100)
        font_weight = st.selectbox("字体粗细", ["normal", "bold"], index=1)
    
    # 姓名位置调整
    st.sidebar.subheader("📍 姓名位置调整")
    col_offset1, col_offset2 = st.sidebar.columns(2)
    with col_offset1:
        text_offset_x = st.slider("水平偏移(mm)", -50.0, 50.0, 0.0, 0.5)
    with col_offset2:
        text_offset_y = st.slider("垂直偏移(mm)", -50.0, 50.0, 0.0, 0.5)
    
    # 字间距调整
    st.sidebar.subheader("🔤 字间距调整")
    char_spacing = st.sidebar.slider("字间距(pt)", -10.0, 20.0, 0.0, 0.5)
    
    # 颜色设置
    font_color = st.sidebar.color_picker("字体颜色", "#000000")
    background_color = st.sidebar.color_picker("背景颜色", "#ffffff")
    border_color = st.sidebar.color_picker("边框颜色", "#cccccc")
    
    # 背景图片设置
    st.sidebar.subheader("🖼️ 背景图片")
    background_image = st.sidebar.file_uploader(
        "上传背景图片",
        type=["jpg", "jpeg", "png", "gif"],
        help="支持 JPG、JPEG、PNG、GIF 格式"
    )
    
    # 其他设置
    st.sidebar.subheader("📝 字体设置")
    
    # 字体选择
    font_family = st.sidebar.selectbox(
        "系统字体", 
        ["Microsoft YaHei", "SimHei", "Arial", "Times New Roman"],
        index=0
    )
    
    # 读取挂载目录中的字体
    FONTS_DIR = "/app/fonts" if os.path.exists("/app/fonts") else "fonts"
    available_fonts = []
    font_files = {}
    
    if os.path.exists(FONTS_DIR):
        try:
            # 扫描字体文件
            for filename in os.listdir(FONTS_DIR):
                if filename.lower().endswith(('.ttf', '.otf', '.ttc')):
                    font_name = os.path.splitext(filename)[0]
                    available_fonts.append(font_name)
                    font_files[font_name] = os.path.join(FONTS_DIR, filename)
        except Exception as e:
            st.warning(f"无法读取字体目录: {e}")
            available_fonts = []
            font_files = {}
    
    # 系统字体选择
    selected_system_font = None
    if available_fonts:
        st.sidebar.subheader("📁 挂载字体")
        selected_system_font = st.sidebar.selectbox(
            "选择挂载字体",
            ["（不使用）"] + available_fonts,
            help="从挂载目录 /app/fonts 中选择字体，用户可预先将字体文件放入 ./fonts 目录"
        )
        if selected_system_font == "（不使用）":
            selected_system_font = None
    
    # 自定义字体上传
    custom_font = st.sidebar.file_uploader(
        "上传自定义字体",
        type=["ttf", "otf"],
        help="支持 TTF、OTF 格式字体文件，如果已选择挂载字体，此选项将被忽略"
    )
    
    # 字体效果设置
    st.sidebar.subheader("✨ 字体效果")
    
    # 支持效果叠加
    text_effects = st.sidebar.multiselect(
        "文字效果（可多选）",
        ["立体", "阴影", "勾边"],
        default=[],
        help="选择多个效果可以叠加使用"
    )
    
    # 阴影效果设置
    if "阴影" in text_effects:
        shadow_offset = st.sidebar.slider("阴影偏移", 0.5, 5.0, 2.0, 0.1)
        shadow_color = st.sidebar.color_picker("阴影颜色", "#888888")
    else:
        shadow_offset = 2.0
        shadow_color = "#888888"
    
    # 勾边效果设置
    if "勾边" in text_effects:
        stroke_width = st.sidebar.slider("勾边宽度", 0.5, 3.0, 1.0, 0.1)
        stroke_color = st.sidebar.color_picker("勾边颜色", "#000000")
    else:
        stroke_width = 1.0
        stroke_color = "#000000"
    
    # 立体效果设置
    if "立体" in text_effects:
        stereo_offset = st.sidebar.slider("立体偏移", 0.5, 3.0, 1.0, 0.1)
        stereo_color = st.sidebar.color_picker("立体颜色", "#666666")
    else:
        stereo_offset = 1.0
        stereo_color = "#666666"
    
    border_width = st.sidebar.slider("边框宽度", 0, 10, 0)
    border_style = st.sidebar.selectbox("边框样式", ["solid", "dashed", "dotted"])
    card_type = st.sidebar.selectbox(
        "座位牌类型",
        ["普通座位牌", "三角立式台卡"],
        index=0,
        help="普通座位牌：直接插入模具；三角立式台卡：需要折叠，前后两面内容相同，包含镜像打印"
    )
    
    # 座位号设置
    st.sidebar.subheader("📌 座位号设置")
    show_seat_number = st.sidebar.checkbox("显示座位号", value=True)
    if show_seat_number:
        col1, col2 = st.sidebar.columns(2)
        with col1:
            seat_number_font_size = st.number_input(
                "座位号字体大小(px)", min_value=5, max_value=30, value=6
            )
        with col2:
            seat_number_color = st.sidebar.color_picker("座位号颜色", "#666666")
    else:
        seat_number_font_size = 6
        seat_number_color = "#666666"
    
    # 三角立式台卡专属设置
    if card_type == '三角立式台卡':
        st.sidebar.subheader("📎 粘贴区设置")
        paste_area_height = st.sidebar.number_input(
            "空白粘贴区高度(mm)", min_value=10.0, max_value=150.0, value=48.5, step=0.5
        )
        paste_area_color = st.sidebar.color_picker("空白粘贴区颜色", "#ffffff")
    else:
        paste_area_height = 48.5
        paste_area_color = "#ffffff"
    
    page_size = st.sidebar.selectbox(
        "打印纸张尺寸", 
        ["A3", "A4", "A5", "A6", "Letter"],
        index=1
    )
    
    # 主内容区域
    st.header("📊 数据导入")
    
    # 数据来源选择
    data_source = st.radio(
        "选择数据来源",
        ["📁 上传文件 / 手动输入", "🗄️ 从座位查询系统导入"],
        horizontal=True,
        key="data_source_radio"
    )
    
    names = []
    name_position_list = []  # [(name, position), ...] 列表，支持重名
    
    if data_source == "🗄️ 从座位查询系统导入":
        # ===== 从座位查询系统导入已识别的数据 =====
        st.subheader("🗄️ 从座位查询系统导入")
        
        finder_data, error = load_attendees_from_finder()
        
        if error:
            st.warning(error)
        else:
            st.success(f"成功读取座位查询系统数据，共 {len(finder_data['attendees'])} 个参会者")
            
            # 会场选择
            venue_ids = list(finder_data['venue_attendees'].keys())
            if len(venue_ids) > 1:
                selected_venue_id = st.selectbox(
                    "选择会场",
                    venue_ids,
                    format_func=lambda vid: finder_data['venue_names'].get(vid, vid)
                )
            else:
                selected_venue_id = venue_ids[0]
            
            selected_attendees = finder_data['venue_attendees'][selected_venue_id]
            
            # 按排分组
            from collections import OrderedDict
            rows = OrderedDict()
            for a in selected_attendees:
                row_label = a.get('row', '未指定')
                if row_label not in rows:
                    rows[row_label] = []
                rows[row_label].append(a)
            
            # 按排序号排序
            def row_sort_key(item):
                label = item[0]
                m = re.match(r'(?:沙发)?第(.+)排', label)
                if m:
                    return cnNums.get(m[1], 999)
                return 999
            
            sorted_rows = sorted(rows.items(), key=row_sort_key)
            
            st.write("请选择要生成座位牌的排和参会者:")
            
            col1, col2, col3 = st.columns([1, 1, 3])
            with col1:
                if st.button("✅ 全选所有", key="finder_select_all"):
                    st.session_state["finder_select_all_rows"] = True
            with col2:
                if st.button("❌ 取消全选", key="finder_deselect_all"):
                    st.session_state["finder_select_all_rows"] = False
            with col3:
                st.caption(f"共 {len(selected_attendees)} 人 | {len(sorted_rows)} 排")
            
            st.write("---")
            
            # 构建完整的参会者列表（含排信息）
            all_attendee_list = []
            attendee_index = 0
            for row_label, row_attendees in sorted_rows:
                for a in row_attendees:
                    all_attendee_list.append({
                        'index': attendee_index,
                        'name': a['name'],
                        'row': a['row'],
                        'seat': a['seat'],
                        'company': a.get('company', ''),
                        'title': a.get('title', '')
                    })
                    attendee_index += 1
            
            # 按排显示
            selected_indices = set()
            
            for row_label, row_attendees in sorted_rows:
                row_count = len(row_attendees)
                row_key = f"finder_row_{row_label}"
                default_expanded = st.session_state.get("finder_select_all_rows", True)
                
                with st.expander(f"{row_label}（{row_count}人）", expanded=default_expanded):
                    row_selected = st.checkbox(
                        "☑ 全选本排",
                        value=st.session_state.get(row_key, True),
                        key=row_key
                    )
                    
                    if row_selected:
                        st.text("  ".join([a['name'] for a in row_attendees[:8]]) + ("..." if len(row_attendees) > 8 else ""))
                        for a in row_attendees:
                            for item in all_attendee_list:
                                if item['name'] == a['name'] and item['row'] == a['row'] and item['seat'] == a['seat']:
                                    selected_indices.add(item['index'])
                    else:
                        # 按行显示人员并允许单独选择
                        cols = st.columns(3)
                        for ci, a in enumerate(row_attendees):
                            with cols[ci % 3]:
                                cb_key = f"finder_p_{a['name']}_{a['row']}_{a['seat']}".replace(" ", "_")
                                is_checked = st.checkbox(
                                    f"{a['name']}（{a['seat']}号）",
                                    value=st.session_state.get(cb_key, False),
                                    key=cb_key
                                )
                                if is_checked:
                                    for item in all_attendee_list:
                                        if item['name'] == a['name'] and item['row'] == a['row'] and item['seat'] == a['seat']:
                                            selected_indices.add(item['index'])
            
            st.write("---")
            
            if selected_indices:
                selected_items = [all_attendee_list[i] for i in sorted(selected_indices)]
                
                # 构建 name_position_list
                name_position_list = []
                names = []
                for item in selected_items:
                    position = f"{item['row']} {item['seat']}号"
                    name_position_list.append((item['name'], position))
                    names.append(item['name'])
                
                data_path_display = os.path.join('data', 'data.json')
                st.info(f"已从 [{data_path_display}](file:///d:/autofill/seat-finder/data.json) 选择 {len(names)} 个参会者生成座位牌")
            else:
                st.warning("请至少选择一个排或参会者")
    else:
        # ===== 原有的文件上传/手动输入逻辑 =====
        # 文件上传
        uploaded_file = st.file_uploader(
            "上传姓名列表文件", 
            type=["csv", "xlsx"],
            help="支持CSV或Excel格式，文件应包含姓名列"
        )
        
        # 座位布局Excel文件上传
        layout_file = st.file_uploader(
            "上传座位布局Excel文件（可选）", 
            type=["xlsx"],
            help="上传包含座位布局的Excel文件，系统会自动提取参会者位置信息"
        )
        
        # 手动输入备选
        manual_names = st.text_area(
            "或手动输入姓名（每行一个）",
            placeholder="张三\n李四\n王五",
            height=100
        )
        
        # 处理座位布局文件
        if layout_file is not None:
            try:
                # 保存上传的文件到临时文件
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    temp_file.write(layout_file.getbuffer())
                    temp_file_path = temp_file.name
                
                # 优先使用座位查找系统的解析逻辑
                name_position_list = parseExcelLayoutFromFinder(temp_file_path)
                
                # 如果座位查找系统没有解析出结果，回退到原有解析方法
                if not name_position_list:
                    print("[回退] 座位查找系统未解析出结果，使用内置解析方法")
                    name_position_list = parseExcelLayout(temp_file_path)
                
                # 清理临时文件
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                
                if name_position_list:
                    # 按姓名分组显示（支持重名）
                    from collections import OrderedDict
                    name_positions = OrderedDict()
                    for n, p in name_position_list:
                        if n not in name_positions:
                            name_positions[n] = []
                        name_positions[n].append(p)
                    
                    # 统计重名数量
                    duplicate_count = sum(1 for positions in name_positions.values() if len(positions) > 1)
                    
                    st.success(f"成功解析座位布局，找到 {len(name_position_list)} 个位置信息（含 {len(name_positions)} 个不重复姓名）")
                    if duplicate_count > 0:
                        st.info(f"其中 {duplicate_count} 个姓名有重复（共 {sum(len(p) for p in name_positions.values() if len(p) > 1)} 条记录）")
                    
                    # 显示提取的位置信息并添加复选框
                    st.write("请选择要生成座位牌的参会者:")
                    
                    # 添加全选/取消全选按钮
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✅ 全选"):
                            # 使用session_state来存储选择状态
                            for idx in range(len(name_position_list)):
                                st.session_state[f"select_attendee_{idx}"] = True
                    with col2:
                        if st.button("❌ 取消全选"):
                            for idx in range(len(name_position_list)):
                                st.session_state[f"select_attendee_{idx}"] = False
                    
                    st.write("---")
                    
                    # 存储用户选择的索引
                    selected_indices = []
                    
                    # 遍历显示所有参会者，带复选框
                    for idx, (n, p) in enumerate(name_position_list):
                        # 为每个参会者创建唯一的复选框key
                        checkbox_key = f"select_attendee_{idx}"
                        # 从session_state获取值，默认选中
                        default_value = st.session_state.get(checkbox_key, True)
                        is_selected = st.checkbox(f"{n}: {p}", value=default_value, key=checkbox_key)
                        if is_selected:
                            selected_indices.append(idx)
                    
                    st.write("---")
                    
                    # 计算选中的参会者
                    if selected_indices:
                        # 只保留选中的参会者
                        selected_name_position_list = [name_position_list[i] for i in selected_indices]
                        # 更新name_position_list为选中的
                        name_position_list = selected_name_position_list
                        # 更新names列表为选中的姓名
                        names = [n for n, p in name_position_list]
                        st.info(f"已选择 {len(names)} 个参会者生成座位牌")
                    else:
                        st.warning("请至少选择一个参会者")
                else:
                    st.warning("未从座位布局文件中提取到位置信息")
            except Exception as e:
                st.error(f"解析座位布局失败: {e}")
                import traceback
                traceback.print_exc()
        
        if uploaded_file is not None:
            # 处理上传的文件
            try:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
                
                # 自动检测姓名列
                name_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in ['name', '姓名', '名字'])]
                
                if name_columns:
                    name_column = st.selectbox("选择姓名列", name_columns)
                    names = df[name_column].dropna().astype(str).tolist()
                else:
                    name_column = st.selectbox("选择姓名列", df.columns)
                    names = df[name_column].dropna().astype(str).tolist()
                
                st.success(f"成功导入 {len(names)} 个姓名")
                
            except Exception as e:
                st.error(f"文件读取错误: {e}")
        
        elif manual_names.strip():
            # 处理手动输入的姓名
            names = [name.strip() for name in manual_names.strip().split('\n') if name.strip()]
            st.success(f"成功导入 {len(names)} 个姓名")
    
    # 实时预览功能（始终显示）
    st.subheader("👀 实时效果预览")
    
    # 当参数变化时自动更新预览
    preview_container = st.container()
    
    with preview_container:
        with st.spinner("正在生成预览..."):
            try:
                # 准备模板变量
                template_vars = {
                    'page_size': page_size,
                    'card_type': card_type,
                    'background_color': background_color,
                    'card_width': card_width,
                    'card_height': card_height,
                    'border_width': border_width,
                    'border_style': border_style,
                    'border_color': border_color,
                    'card_background': background_color,
                    'font_family': font_family,
                    'font_size': font_size,
                    'font_weight': font_weight,
                    'font_color': font_color,
                    'background_image': background_image,
                    'custom_font': custom_font,
                    'selected_system_font': selected_system_font,
                    'font_files': font_files,
                    'text_effects': text_effects,
                    'shadow_offset': shadow_offset,
                    'shadow_color': shadow_color,
                    'stroke_width': stroke_width,
                    'stroke_color': stroke_color,
                    'stereo_offset': stereo_offset,
                    'stereo_color': stereo_color,
                    'text_offset_x': text_offset_x,
                    'text_offset_y': text_offset_y,
                    'char_spacing': char_spacing,
                    'name_position_list': name_position_list,
                    'paste_area_color': paste_area_color,
                    'paste_area_height': paste_area_height,
                    'show_seat_number': show_seat_number,
                    'seat_number_font_size': seat_number_font_size,
                    'seat_number_color': seat_number_color
                }
                
                # 为预览保存上传的文件为临时文件
                preview_temp_custom_font_path = None
                preview_temp_custom_font_name = None
                preview_temp_background_image_path = None
                preview_temp_files_to_cleanup = []
                
                if custom_font:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.' + custom_font.name.split('.')[-1]) as temp_file:
                        temp_file.write(custom_font.getbuffer())
                        preview_temp_custom_font_path = temp_file.name
                        preview_temp_custom_font_name = os.path.splitext(os.path.basename(custom_font.name))[0]
                        preview_temp_files_to_cleanup.append(preview_temp_custom_font_path)
                
                if background_image:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.' + background_image.name.split('.')[-1]) as temp_file:
                        temp_file.write(background_image.getbuffer())
                        preview_temp_background_image_path = temp_file.name
                        preview_temp_files_to_cleanup.append(preview_temp_background_image_path)
                
                # 生成预览PDF（只生成第一个姓名，使用足够大的页面）
                preview_names = names[:1] if names else ["预览"]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                preview_filename = f"preview_{timestamp}.pdf"
                preview_path = os.path.join("output", preview_filename)
                
                # 为预览创建保持比例的座位牌，确保文字清晰可见
                temp_vars = template_vars.copy()
                
                # 移除UploadedFile对象，添加临时文件路径
                temp_vars['custom_font'] = None
                temp_vars['background_image'] = None
                temp_vars['custom_font_path'] = preview_temp_custom_font_path
                temp_vars['custom_font_name'] = preview_temp_custom_font_name
                temp_vars['background_image_path'] = preview_temp_background_image_path
                
                # 保持用户设置的宽高比
                original_width = temp_vars.get('card_width', 90)
                original_height = temp_vars.get('card_height', 55)
                
                # 对于预览，只考虑单个座位牌的比例，不包括上下空白
                aspect_ratio = original_width / original_height
                
                # 为预览设置合适的基础宽度，确保在浏览器中显示合适
                base_preview_width = 120  # 预览座位牌基础宽度（更小巧，只作为效果预览）
                preview_card_width = base_preview_width
                preview_card_height = preview_card_width / aspect_ratio
                
                # 计算字体缩放比例（基于单个座位牌）
                font_scale = preview_card_width / original_width
                
                temp_vars['card_width'] = preview_card_width
                temp_vars['card_height'] = preview_card_height
                # 按比例缩放字体大小
                temp_vars['font_size'] = int(temp_vars.get('font_size', 48) * font_scale)
                
                # 预览时，如果是三角立式台卡，暂时改为普通座位牌以只显示单个座位牌
                original_card_type = temp_vars.get('card_type')
                if temp_vars.get('card_type') == '三角立式台卡':
                    temp_vars['card_type'] = '普通座位牌'  # 预览时使用普通座位牌
                    temp_vars['page_width'] = preview_card_width  # 左右不留空白
                    temp_vars['page_height'] = preview_card_height
                else:
                    # 普通座位牌：使用自定义页面尺寸
                    temp_vars['page_width'] = preview_card_width  # 左右不留空白
                    temp_vars['page_height'] = preview_card_height
                
                try:
                    generate_pdf(preview_names, temp_vars, preview_path)
                    
                    # 恢复原始的座位牌类型，确保生成PDF时使用正确的类型
                    temp_vars['card_type'] = original_card_type
                    
                    # 显示预览
                    st.success("✅ 预览生成完成！")
                    
                    # 提供预览文件
                    with open(preview_path, "rb") as f:
                        preview_data = f.read()
                    
                    st.download_button(
                        label="📄 查看预览PDF",
                        data=preview_data,
                        file_name=preview_filename,
                        mime="application/pdf"
                    )
                    
                    # 显示预览图片（使用PDF的第一页）
                    try:
                        from pdf2image import convert_from_path
                        # 根据操作系统指定 Poppler 路径
                        if sys.platform == 'win32':
                            poppler_path = r"C:\Program Files\poppler\Library\bin"
                            images = convert_from_path(preview_path, first_page=1, last_page=1, poppler_path=poppler_path, dpi=300)
                        else:
                            # Linux/Docker 环境下 poppler 在系统 PATH 中
                            images = convert_from_path(preview_path, first_page=1, last_page=1, dpi=300)
                        if images:
                            import io
                            buf = io.BytesIO()
                            images[0].save(buf, format='PNG')
                            buf.seek(0)
                            # 让预览图片自动跟随浏览器页面缩放
                            st.image(buf, caption="座位牌预览", use_container_width=True)
                    except Exception as e:
                        st.info(f"预览图片生成失败：{str(e)}，您可以下载PDF查看效果")
                except Exception as e:
                    st.error(f"预览生成失败: {e}")
                    import traceback
                    st.error(f"详细信息: {traceback.format_exc()}")
                finally:
                    # 清理临时文件
                    for temp_file in preview_temp_files_to_cleanup:
                        try:
                            if os.path.exists(temp_file):
                                os.unlink(temp_file)
                        except:
                            pass
            except Exception as e:
                st.error(f"预览准备失败: {e}")
                import traceback
                st.error(f"详细信息: {traceback.format_exc()}")
    
    # 显示姓名预览
    if names:
        st.subheader("👥 姓名预览")
        st.write(f"共 {len(names)} 个姓名:")
        
        # 分列显示
        cols = st.columns(4)
        for i, name in enumerate(names):
            with cols[i % 4]:
                st.text(f"{i+1}. {name}")
        
        # 生成按钮
        st.header("🖨️ 生成座位牌")
        
        # 生成PDF按钮
        if st.button("🚀 生成PDF文件", type="primary"):
            with st.spinner("正在生成座位牌..."):
                try:
                    # 准备模板变量
                    template_vars = {
                        'page_size': page_size,
                        'card_type': card_type,
                        'background_color': background_color,
                        'card_width': card_width,
                        'card_height': card_height,
                        'border_width': border_width,
                        'border_style': border_style,
                        'border_color': border_color,
                        'card_background': background_color,
                        'font_family': font_family,
                        'font_size': font_size,
                        'font_weight': font_weight,
                        'font_color': font_color,
                        'background_image': background_image,
                        'custom_font': custom_font,
                        'selected_system_font': selected_system_font,
                        'font_files': font_files,
                        'text_effects': text_effects,
                        'shadow_offset': shadow_offset,
                        'shadow_color': shadow_color,
                        'stroke_width': stroke_width,
                        'stroke_color': stroke_color,
                        'stereo_offset': stereo_offset,
                        'stereo_color': stereo_color,
                        'text_offset_x': text_offset_x,
                        'text_offset_y': text_offset_y,
                        'char_spacing': char_spacing,
                        'name_position_list': name_position_list,
                        'paste_area_color': paste_area_color,
                        'paste_area_height': paste_area_height,
                        'show_seat_number': show_seat_number,
                        'seat_number_font_size': seat_number_font_size,
                        'seat_number_color': seat_number_color
                    }
                    
                    # 在调用generate_pdf之前，先保存上传的文件为临时文件
                    temp_custom_font_path = None
                    temp_custom_font_name = None
                    temp_background_image_path = None
                    temp_files_to_cleanup = []
                    
                    if custom_font:
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.' + custom_font.name.split('.')[-1]) as temp_file:
                            temp_file.write(custom_font.getbuffer())
                            temp_custom_font_path = temp_file.name
                            temp_custom_font_name = os.path.splitext(os.path.basename(custom_font.name))[0]
                            temp_files_to_cleanup.append(temp_custom_font_path)
                    
                    if background_image:
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.' + background_image.name.split('.')[-1]) as temp_file:
                            temp_file.write(background_image.getbuffer())
                            temp_background_image_path = temp_file.name
                            temp_files_to_cleanup.append(temp_background_image_path)
                    
                    # 更新模板变量，只传递临时文件路径而不是UploadedFile对象
                    template_vars_safe = template_vars.copy()
                    template_vars_safe['custom_font'] = None  # 移除UploadedFile对象
                    template_vars_safe['background_image'] = None  # 移除UploadedFile对象
                    template_vars_safe['custom_font_path'] = temp_custom_font_path
                    template_vars_safe['custom_font_name'] = temp_custom_font_name
                    template_vars_safe['background_image_path'] = temp_background_image_path
                    
                    # 生成输出文件名
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename = f"namecards_{timestamp}.pdf"
                    output_path = os.path.join("output", output_filename)
                    
                    try:
                        # 为每个姓名生成PDF
                        generate_pdf(names, template_vars_safe, output_path)
                        
                        st.success(f"✅ 座位牌生成完成！共生成 {len(names)} 个座位牌")
                        
                        # 提供下载链接
                        with open(output_path, "rb") as f:
                            pdf_data = f.read()
                        
                        st.download_button(
                            label="📥 下载PDF文件",
                            data=pdf_data,
                            file_name=output_filename,
                            mime="application/pdf"
                        )
                    except Exception as e:
                        st.error(f"生成失败: {e}")
                        import traceback
                        st.error(f"详细信息: {traceback.format_exc()}")
                    finally:
                        # 清理临时文件
                        for temp_file in temp_files_to_cleanup:
                            try:
                                if os.path.exists(temp_file):
                                    os.unlink(temp_file)
                            except:
                                pass
                except Exception as e:
                    st.error(f"准备生成失败: {e}")
                    import traceback
                    st.error(f"详细信息: {traceback.format_exc()}")
    
    else:
        st.info("👆 请先上传姓名列表文件或手动输入姓名")
    
    # 使用说明
    st.sidebar.header("💡 使用说明")
    st.sidebar.markdown("""
    1. 上传CSV/Excel文件或手动输入姓名
    2. 调整样式设置（大小、颜色、字体等）
    3. 点击"生成PDF文件"按钮
    4. 下载并打印生成的PDF
    """)

if __name__ == "__main__":
    main()